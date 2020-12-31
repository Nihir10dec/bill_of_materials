from os import write
from openpyxl import load_workbook, workbook
from collections import defaultdict

items = defaultdict(list) # creating a dictionary with default value as list to directly append new items to it

def read_from_excel(fname):
    workbook = load_workbook(fname)
    sheet = workbook["Source"] # loading the source sheet to read data from it
    
    for row in sheet.iter_rows(min_row=2,max_row=12,min_col=1,max_col=5,values_only=True):
        item ={}
        item["name"] = row[0]        
        item["level"] = int(row[1][-1])
        item["raw material"] = row[2]
        item["quantity"] = row[3]
        item["unit"] = row[4]
        # creting dictionary object with all the details and appending to main ITEMS dictionary created above
        items[row[0]].append(item)
    print("Read from Excel")

# this function destructures the dictionary while keeping only level 1 objects for the main items and makes new record for the level 1 items made of level 2 objects and similary for level 2 and 3.
def organise():
    prev_raw1,prev_raw2 = {},{}
    list1 =[]
    for item in list(items):
        for i in items[item]:
            if i["level"] == 1 and i["name"] == item:
                list1.append(i)
                prev_raw1 = i
            elif i["level"] == 2:
                items[prev_raw1["raw material"]].append(i)
                prev_raw2 = i
            elif i["level"] ==3:
                items[prev_raw2["raw material"]].append(i)
            items[item] = list1
        list1=[]
    print("Organised the dictionary to write back to excel")


def write_to_excel(items_to_write , fname):
    workbook = load_workbook(fname)
    for item in items_to_write:
        if(item != 'Fan'): # using the already created fan sheet as template for others 
            fan_sheet = workbook["Fan"]
            if item in workbook.sheetnames:
                workbook.remove(workbook[item])
                
            new_sheet = workbook.copy_worksheet(fan_sheet)
            new_sheet.title = item 
            new_sheet["B3"].value = item
                
            x = 7
            for raw_material in items_to_write[item]:         
                # replacing the values copied from the sheet with it's original values.       
                new_sheet["B" + str(x)].value = raw_material["raw material"]
                new_sheet["C" + str(x)].value = raw_material["quantity"]
                new_sheet["D" + str(x)].value = raw_material["unit"]
                x+=1
                    
            if(len(items_to_write[item]) <3):
                # removing the extra values and making it as blank
                # we can also delete the cells (sheet.delete_cols(idx) ) but to keep all sheets as same as possible have made the value as empty.
                while(True):
                    new_sheet["A" + str(x)].value = ""
                    new_sheet["B" + str(x)].value = ""
                    new_sheet["C" + str(x)].value = ""
                    new_sheet["D" + str(x)].value = ""
                    x+=1
                    if(x>=10):
                        break
            else:
                new_sheet["A" + str(x)].value = "End of RM"

    workbook.save(fname) # saving the workbook to reflect the changes
    print("Wrote back to Excel")


read_from_excel("BOM file for Data processing (1) (1).xlsx")
organise()
write_to_excel(dict(items) , "BOM file for Data processing (1) (1).xlsx")
